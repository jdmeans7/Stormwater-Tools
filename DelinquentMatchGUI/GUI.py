import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
import openpyxl
import time
import threading


root = tk.Tk()

o = tk.IntVar()
df = tk.StringVar()
pf = tk.StringVar()
progressvar = tk.DoubleVar()
tk.Label(root, text="Load spreadsheets using the two load buttons, "
                    "select the configuration you need, then click execute.").pack(anchor=tk.N)
progressbar = ttk.Progressbar(root, variable=progressvar)
progressbar.pack(expand=1)
root.title("Delinquent Matcher")


Options = [("No repeating information, only parcel added", 1),
           ("No repeating information, parcel and location added", 2),
           ("No repeating information, parcel, location, and sq feet added", 3),
           ("Repeating information, parcel, location, and sq feet added", 4)]


def loadDelinquent():
    global df
    df = filedialog.askopenfilename()
    print(df)


def loadParcels():
    global pf
    pf = filedialog.askopenfilename()
    print(pf)


tk.Button(root, text="Load delinquent file", command=loadDelinquent).pack(anchor=tk.W)
tk.Button(root, text="Load parcel file", command=loadParcels).pack(anchor=tk.W)


def ShowChoice():
    return o.get()


for option, val in Options:
    tk.Radiobutton(root,
                   text=option,
                   padx=20,
                   variable=o,
                   command=ShowChoice,
                   value=val).pack(anchor=tk.W)


def loadin():
    global df
    global pf
    delinquentwb = openpyxl.load_workbook(df)
    parcelwb = openpyxl.load_workbook(pf)
    delinquentsheet = delinquentwb.active
    parcelsheet = parcelwb.active
    progressbar.configure(maximum=delinquentsheet.max_row)
    return delinquentsheet, parcelsheet, delinquentwb


def Execute():
    delinquentsheet, parcelsheet, delinquentwb = loadin()
    acctnums = list()
    for cell in tuple(delinquentsheet.columns)[1]:
        acctnums.append(cell.value)

    acctprop = dict()
    proploc = dict()
    parcelsqft = dict()

    if ShowChoice() == 1:
        for i in range(1, parcelsheet.max_row, 1):
            props = list()
            props.append(parcelsheet.cell(row=i, column=16).value)
            acct = parcelsheet.cell(row=i, column=1).value
            if acct not in acctprop:
                while parcelsheet.cell(row=i, column=1).value == parcelsheet.cell(row=i + 1, column=1).value:
                    props.append(parcelsheet.cell(row=i + 1, column=16).value)
                    i += 1
                acctprop[acct] = props
    if ShowChoice() == 2:
        for i in range(1, parcelsheet.max_row, 1):
            proploc[parcelsheet.cell(row=i, column=16).value] = parcelsheet.cell(row=i, column=17).value
            props = list()
            props.append(parcelsheet.cell(row=i, column=16).value)
            acct = parcelsheet.cell(row=i, column=1).value
            if acct not in acctprop:
                while parcelsheet.cell(row=i, column=1).value == parcelsheet.cell(row=i + 1, column=1).value:
                    props.append(parcelsheet.cell(row=i + 1, column=16).value)
                    i += 1
                acctprop[acct] = props
    elif ShowChoice() == 3 or ShowChoice() == 4:
        for i in range(1, parcelsheet.max_row, 1):
            proploc[parcelsheet.cell(row=i, column=16).value] = parcelsheet.cell(row=i, column=17).value
            parcelsqft[parcelsheet.cell(row=i, column=16).value] = parcelsheet.cell(row=i, column=19).value
            props = list()
            props.append(parcelsheet.cell(row=i, column=16).value)
            acct = parcelsheet.cell(row=i, column=1).value
            if acct not in acctprop:
                while parcelsheet.cell(row=i, column=1).value == parcelsheet.cell(row=i + 1, column=1).value:
                    props.append(parcelsheet.cell(row=i + 1, column=16).value)
                    i += 1
                acctprop[acct] = props

    i = 1
    while i < delinquentsheet.max_row:
        progressbar.step(1)
        time.sleep(0.001)
        root.update_idletasks()
        acct = delinquentsheet.cell(row=i, column=2).value
        if acct in acctprop:
            count = 0
            if len(acctprop[acct]) > 1:
                properties = acctprop[delinquentsheet.cell(row=i, column=2).value]
                for p in properties:
                    if count == 0:
                        delinquentsheet.cell(row=i, column=12).value = p
                        if ShowChoice() == 2 or ShowChoice() == 3 or ShowChoice() == 4:
                            delinquentsheet.cell(row=i, column=13).value = proploc[p]  # add location
                        if ShowChoice() == 3 or ShowChoice() == 4:
                            delinquentsheet.cell(row=i, column=14).value = parcelsqft[p]  # add sqft
                        count += 1
                    else:
                        delinquentsheet.insert_rows(i + 1)
                        if ShowChoice() == 4:
                            for a in range(1, 12):
                                # if a == 10 or a == 11:
                                #    delinquentsheet.cell(row=i+1, column=a).value = 0
                                # else:
                                #    delinquentsheet.cell(row=i+1, column=a).value = delinquentsheet.cell(row=i, column=a).value
                                delinquentsheet.cell(row=i + 1, column=a).value = delinquentsheet.cell(row=i,
                                                                                                       column=a).value
                        delinquentsheet.cell(row=i + 1, column=12).value = p
                        if ShowChoice() == 2 or ShowChoice() == 3 or ShowChoice() == 4:
                            delinquentsheet.cell(row=i + 1, column=13).value = proploc[p]  # add location
                        if ShowChoice() == 3 or ShowChoice() == 4:
                            delinquentsheet.cell(row=i + 1, column=14).value = parcelsqft[p]  # add sqft
                        i += 1
                i += 1
            else:
                delinquentsheet.cell(row=i, column=12).value = acctprop[acct][0]
                if ShowChoice() == 2 or ShowChoice() == 3 or ShowChoice() == 4:
                    delinquentsheet.cell(row=i, column=13).value = proploc[acctprop[acct][0]]  # add location
                if ShowChoice() == 3 or ShowChoice() == 4:
                    delinquentsheet.cell(row=i, column=14).value = parcelsqft[acctprop[acct][0]]  # add sqft
                i += 1
        else:
            i += 1

    f = filedialog.asksaveasfilename(defaultextension=".xlsx")
    delinquentwb.save(f)
    tk.Label(root, text="Finished!").pack()


tk.Button(root, text="Execute", command=Execute).pack()


root.mainloop()
