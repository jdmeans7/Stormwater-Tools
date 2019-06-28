import openpyxl


def main():
    delinquentsheet, parcelsheet = loadWBs("UBCOLWS4.5.19", "AllParcelsNoStops4.5.19")
    acctnums = list()
    for cell in tuple(delinquentsheet.columns)[1]:
        acctnums.append(cell.value)

    acctprop = dict()
    proploc = dict()
    parcelsqft = dict()

    i = 1
    print("Select the option you wish:")
    print("1. No repeating information, only parcel added")
    print("2. No repeating information, parcel and location added")
    print("3. No repeating information, parcel, location, and sq feet added")
    print("4. Repeating information, parcel, location, and sq feet added")
    option = input(">")
    load(option)
    while i < delinquentsheet.max_row:
        acct = delinquentsheet.cell(row=i, column=2).value
        if acct in acctprop:
            count = 0
            if len(acctprop[acct]) > 1:
                properties = acctprop[delinquentsheet.cell(row=i, column=2).value]
                for p in properties:
                    if count == 0:
                        delinquentsheet.cell(row=i, column=12).value = p
                        if option == 2 or option == 3 or option == 4:
                            delinquentsheet.cell(row=i, column=13).value = proploc[p]  # add location
                        if option == 3 or option == 4:
                            delinquentsheet.cell(row=i, column=14).value = parcelsqft[p]  # add sqft
                        count += 1
                    else:
                        delinquentsheet.insert_rows(i + 1)
                        if option == 4:
                            for a in range(1, 12):
                                # if a == 10 or a == 11:
                                #    delinquentsheet.cell(row=i+1, column=a).value = 0
                                # else:
                                #    delinquentsheet.cell(row=i+1, column=a).value = delinquentsheet.cell(row=i, column=a).value
                                delinquentsheet.cell(row=i + 1, column=a).value = delinquentsheet.cell(row=i,
                                                                                                       column=a).value
                        delinquentsheet.cell(row=i + 1, column=12).value = p
                        if option == 2 or option == 3 or option == 4:
                            delinquentsheet.cell(row=i + 1, column=13).value = proploc[p]  # add location
                        if option == 3 or option == 4:
                            delinquentsheet.cell(row=i + 1, column=14).value = parcelsqft[p]  # add sqft
                        i += 1
                i += 1
            else:
                delinquentsheet.cell(row=i, column=12).value = acctprop[acct][0]
                if option == 2 or option == 3 or option == 4:
                    delinquentsheet.cell(row=i, column=13).value = proploc[acctprop[acct][0]]  # add location
                if option == 3 or option == 4:
                    delinquentsheet.cell(row=i, column=14).value = parcelsqft[acctprop[acct][0]]  # add sqft
                i += 1
        else:
            i += 1

    delinquentwb.save("TEST.xlsx")


def load(option):
    if option == 1:
        for i in range(1, parcelsheet.max_row, 1):
            props = list()
            props.append(parcelsheet.cell(row=i, column=16).value)
            acct = parcelsheet.cell(row=i, column=1).value
            if acct not in acctprop:
                while parcelsheet.cell(row=i, column=1).value == parcelsheet.cell(row=i + 1, column=1).value:
                    props.append(parcelsheet.cell(row=i + 1, column=16).value)
                    i += 1
                acctprop[acct] = props
    elif option == 2:
        for i in range(1, parcelsheet.max_row, 1):
            proploc[parcelsheet.cell(row=i, column=16).value] = parcelsheet.cell(row=i, column=17).value
            props = list()
            props.append(parcelsheet.cell(row=i, column=16).value)
            acct = parcelsheet.cell(row=i, column=1).value
            if acct not in acctprop:
                while parcelsheet.cell(row=i, column=1).value == parcelsheet.cell(row=i + 1, column=1).value:
                    props.append(parcelsheet.cell(row=i + 1, column=16).value)
                    i += 1
                acctprop[acct] = props
    elif option == 3 or option == 4:
        for i in range(1, parcelsheet.max_row, 1):
            proploc[parcelsheet.cell(row=i, column=16).value] = parcelsheet.cell(row=i, column=17).value
            parcelsqft[parcelsheet.cell(row=i, column=16).value] = parcelsheet.cell(row=i, column=19).value
            props = list()
            props.append(parcelsheet.cell(row=i, column=16).value)
            acct = parcelsheet.cell(row=i, column=1).value
            if acct not in acctprop:
                while parcelsheet.cell(row=i, column=1).value == parcelsheet.cell(row=i + 1, column=1).value:
                    props.append(parcelsheet.cell(row=i + 1, column=16).value)
                    i += 1
                acctprop[acct] = props


def loadWBs(delinquentWB, parcelWB):
    delinquentwb = openpyxl.load_workbook(delinquentWB)
    parcelwb = openpyxl.load_workbook(parcelWB)
    delinquentsheet = delinquentwb.get_active_sheet()
    parcelsheet = parcelwb.get_active_sheet()
    return delinquentsheet, parcelsheet





